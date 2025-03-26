import os
import json
import tempfile
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font
from openpyxl.cell.rich_text import TextBlock, CellRichText
from openpyxl.cell.text import InlineFont
from flask import Flask, request, jsonify, send_file

from flask_cors import CORS

app = Flask(__name__)

CORS(app)
# 加载默认配置
DEFAULT_CONFIG_PATH = 'config.json'
try:
    with open(DEFAULT_CONFIG_PATH, 'r', encoding='utf-8') as f:
        DEFAULT_CONFIG = json.load(f)
        print("默认配置加载成功")
        print(DEFAULT_CONFIG)
except Exception as e:
    print(f"无法加载默认配置: {e}")
    DEFAULT_CONFIG = {}


def load_data(filepath):
    """读取并合并所有sheet的数据"""
    excel_file = pd.ExcelFile(filepath)
    all_dfs = []
    
    # 遍历所有sheet并读取数据
    for sheet_name in excel_file.sheet_names:
        print(f"处理表格: {sheet_name}")
        df = excel_file.parse(sheet_name, skiprows=1)  
        all_dfs.append(df)
    
    # 合并所有数据
    return pd.concat(all_dfs, ignore_index=True)


def calculate_score(row, config):
    """根据配置计算岗位评分"""
    score = 0

    # 1. 专业适配度评分
    professional_score = 0
    for item in config.get('professional', []):
        if item['keyword'] in str(row['本科专业\n名称及代码']):  # 确保转换为字符串
            professional_score = max(professional_score, item['weight'])
    
    professional_score  = config['weight']['适配度'] * professional_score  

    score += professional_score

    # 2. 城市竞争度评分
    city_score = 0
    city_weights = config.get('location',{})["city_weights"]
    for item in city_weights:
        if row['考区'] == item['city']:
            city_score = item['weight']
            break
    
    city_score = config['weight']['城市竞争度'] * city_score

    score += city_score

    # 3. 应届生评分
    graduate_score = 0
    if row['是否限应届毕业生报考'] == '否':
        graduate_score = 73  # 默认权重
    elif row['是否限应届毕业生报考'] == '2025届高校毕业生':
        graduate_score = 100
    elif row['是否限应届毕业生报考'] == '应届毕业生':
        graduate_score = 90
    
    graduate_score = config['weight']['应届生'] * graduate_score
    score += graduate_score

    # 4. 学历评分
    education_score = 0
    education_weights = config.get('education', {}).get('education_weights',[])
    for item in education_weights:
        if row['学历'] == item['value']:
            education_score = item['weight']
            break
    
    education_score = config['weight']['学历'] * education_score
    score += education_score
    

    
    # 5. 发展度评估（岗位层级）
    development_score = 70  # 默认分数
    if '乡镇' in str(row['职位简介']): 
        development_score = 70
    elif '镇' in str(row['招考单位']):
        development_score = 70
    elif '局' in str(row['招考单位']):
        development_score = 93
    elif '政府' in str(row['招考单位']):
        development_score = 95
    elif '院' in str(row['招考单位']):
        development_score = 90
    elif '所' in str(row['招考单位']):
        development_score = 85
    
    development_score = config['weight']['发展度'] * development_score

    score += development_score
    

    
    # 6. 就职地评分
    location_score = 0  
    location_prefs = config.get('location', {}).get('location_preferences', {})
    for location_name, weight in location_prefs.items():
        if location_name in str(row['招考单位']):
            location_score = weight
            break
    
    location_score = config['weight']['就职地'] * location_score

    score += location_score
    


    return round(score, 1), round(professional_score, 1), round(city_score, 1), round(graduate_score, 1), round(education_score, 1), round(development_score, 1), round(location_score, 1)


def highlight_keyword_in_cell(cell, keyword, font_color):
    """修改单元格内指定关键词的颜色"""
    original_text = str(cell.value)  # 确保转换为字符串
    target_index = original_text.find(keyword)
    
    if target_index == -1:
        return  # 关键词不存在时跳过
    
    # 分割文本
    before = original_text[:target_index]
    keyword_part = original_text[target_index:target_index+len(keyword)]
    after = original_text[target_index+len(keyword):]

    # 构建富文本块
    blocks = []
    if before:
        blocks.append(TextBlock(InlineFont(), before))
    blocks.append(TextBlock(InlineFont(color=font_color), keyword_part))
    if after:
        blocks.append(TextBlock(InlineFont(), after))

    # 重新赋值给单元格
    cell.value = CellRichText(*blocks)


def generate_report(df, config, output_path):
    """生成带格式的Excel报告"""
    writer = pd.ExcelWriter(output_path, engine='openpyxl')
    df.to_excel(writer, index=False)
    
    # 设置颜色标注规则
    wb = writer.book
    ws = wb.active
    
    # 定义填充颜色
    red_fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
    yellow_fill = PatternFill(start_color='FFFFE0', end_color='FFFFE0', fill_type='solid')
    green_fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
    gy_fill = PatternFill(start_color='BCE094', end_color='BCE094', fill_type='solid')
    
    # 定义字体颜色
    red_font = Font(color='9C0006')
    yellow_font = Font(color='9C5700')
    green_font = Font(color='006100')
    
    # 获取列的索引
    score_column_index = df.columns.get_loc('岗位总评') + 1
    yingjie_index = df.columns.get_loc('是否限应届毕业生报考') + 1
    major_index = df.columns.get_loc('本科专业\n名称及代码') + 1
    
    # 从配置获取阈值
    excellent_threshold = config.get('thresholds', {}).get('优秀', 87)
    good_threshold = config.get('thresholds', {}).get('良好', 78)
    
    for row in ws.iter_rows(min_row=2, max_col=len(df.columns), max_row=len(df) + 1):
        # 获取单元格
        score_cell = row[score_column_index - 1] 
        score_value = score_cell.value
        
        yj_cell = row[yingjie_index - 1]
        yj_value = yj_cell.value
        
        mj_cell = row[major_index - 1]
        
        # 高亮专业关键词
        for item in config.get('professional', []):
            keyword = item['keyword']
            color = item.get('color', '9C0006')
            # 将十六进制颜色转换为openpyxl格式
            color_code = color if len(color) == 6 else '000000'  # 默认黑色
            highlight_keyword_in_cell(mj_cell, keyword, color_code)
        
        # 应届生单元格格式
        if yj_value == "否":
            yj_cell.fill = red_fill
            yj_cell.font = red_font
        elif yj_value == "2025届高校毕业生":
            yj_cell.fill = green_fill
            yj_cell.font = green_font
        elif yj_value == "应届毕业生":
            yj_cell.fill = gy_fill
            yj_cell.font = green_font
        
        # 评分单元格格式
        if score_value is not None:
            if score_value >= excellent_threshold:
                score_cell.fill = green_fill
                score_cell.font = green_font
            elif score_value >= good_threshold:
                score_cell.fill = gy_fill
                score_cell.font = green_font
            else:
                score_cell.fill = red_fill
                score_cell.font = red_font
    
    writer.close()
    print(f"报告已生成: {output_path}")
    return output_path


def filter_positions(df, config):
    """根据配置筛选职位"""
    # 提取所需要的筛选条件
    education_requirements = config.get('education', {}).get('education_requirements', [])
    print(education_requirements)
    target_cities = config.get('location', {}).get('target_cities', [])
    print(target_cities)
    professional_keywords = [item['keyword'] for item in config.get('professional', [])]
    print(professional_keywords)
    excluded_certificates = config.get('certificate', [])
    print(excluded_certificates)
    
    # 应用筛选条件
    filtered_df = df
    
    # 学历筛选
    if education_requirements:
        filtered_df = filtered_df[filtered_df['学历'].isin(education_requirements)]
    
    # 城市筛选
    if target_cities:
        filtered_df = filtered_df[filtered_df['考区'].isin(target_cities)]
    
    # 专业筛选
    if professional_keywords:
        filtered_df = filtered_df[
            filtered_df['本科专业\n名称及代码'].str.contains('|'.join(professional_keywords), na=False)
        ]
    
    # 排除证书
    if excluded_certificates:
        filtered_df = filtered_df[
            ~filtered_df['其他要求'].fillna('').str.contains('|'.join(excluded_certificates), na=False)
        ]
    
    return filtered_df


def process_data(input_file_path, user_config=None):
    """处理数据并返回结果DataFrame和使用的配置"""
    # 使用用户配置或默认配置
    config = DEFAULT_CONFIG.copy()
    if user_config:
        config.update(user_config)
    
    # 加载数据
    raw_df = load_data(input_file_path)
    
    # 筛选职位
    filtered_df = filter_positions(raw_df, config)
    
    # 计算岗位评分
    filtered_df[['岗位总评','专业适配度得分','城市竞争度得分','应届生得分','学历得分','发展度评估得分（乡镇或市直等）','期望就职地得分']] = filtered_df.apply(
        lambda row: calculate_score(row, config), 
        axis=1,
        result_type='expand'  # 关键修复：展开元组为多列
    )
    
    # 排序结果
    result_df = filtered_df.sort_values(by='岗位总评', ascending=False)
    
    return result_df, config


@app.route('/api/config/default', methods=['GET'])
def get_default_config():
    """获取默认配置"""
    return jsonify(DEFAULT_CONFIG)


@app.route('/api/process', methods=['POST'])
def api_process():
    """处理上传的Excel文件"""
    # 检查是否上传了文件
    if 'file' not in request.files:
        return jsonify({"error": "没有上传文件"}), 400
    
    file = request.files['file']
    if file.filename == '':
        return jsonify({"error": "未选择文件"}), 400
    
    # 检查配置参数
    user_config = None
    if 'config' in request.form:
        try:
            user_config = json.loads(request.form['config'])
            print("Received user config:")
            print(user_config)
        except json.JSONDecodeError:
            return jsonify({"error": "配置格式错误"}), 400
    
    # 保存上传的文件
    temp_dir = tempfile.mkdtemp()
    input_path = os.path.join(temp_dir, file.filename)
    file.save(input_path)
    
    try:
        # 处理数据
        result_df, config = process_data(input_path, user_config)
        
        # 生成报告
        output_path = os.path.join(temp_dir, '岗位筛选报告.xlsx')
        report_path = generate_report(result_df, config, output_path)
        
        # 返回生成的报告
        # 检查文件是否存在
        if not os.path.exists(report_path):
            raise FileNotFoundError("生成的报告文件不存在")

        return send_file(
            report_path,
            as_attachment=True,
            download_name='岗位筛选报告.xlsx',
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
    
    except Exception as e:
        return jsonify({"error": str(e)}), 500
    finally:
        # 清理临时文件
        import shutil
        shutil.rmtree(temp_dir, ignore_errors=True)


if __name__ == '__main__':
    app.run(debug=True, host='0.0.0.0', port=5000)
