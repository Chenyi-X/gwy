import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.styles import Font
from openpyxl.cell.rich_text import TextBlock, CellRichText  
from openpyxl.cell.text import InlineFont
from flask import Flask, request, jsonify, send_file
import os
import tempfile
import json

from flask_cors import CORS


app = Flask(__name__)

CORS(app)

# 全部可配置选项
DEFAULT_CONFIG = {
    # 用户基本筛选条件
    '专业关键词': ['软件工程', '计算机类', '不限'],
    '学历要求': ['大专以上', '本科以上'],
    '是否有基层经验': ['否', ''],
    '应届生': True,
    '排除证书': ['法律职业资格', '会计从业', '中共党员'],
    '意向城市': ['佛山', '云浮', '肇庆'],
    '接受乡镇岗': True,
    '优先级权重': {'适配度': 0.15, '城市竞争度': 0.23, '发展度': 0.25, '应届生': 0.18, '学历': 0.07, '就职地': 0.12},
    
    # 专业适配度评分
    '专业适配度': {
        '软件工程': 100,
        '计算机类': 93,
        '不限': 75
    },
    
    # 城市竞争度评分
    '城市竞争度': {
        '佛山': 70, 
        '肇庆': 85, 
        '云浮': 90
    },
    
    # 应届生状态评分
    '应届生状态': {
        '否': 73,
        '2025届高校毕业生': 100,
        '应届毕业生': 90
    },
    
    # 学历权重评分
    '学历评分': {
        '大专以上': 70,
        '本科以上': 100
    },
    
    # 岗位层级评分（发展度）
    '岗位层级评分': {
        '乡镇': 70,
        '局': 95,
        '政府': 95,
        '院': 90,
        '所': 85,
        '其他': 70
    },
    
    # 就职地区评分
    '就职地区评分': {
        '郁南': 92,
        '新兴': 77,
        '罗定': 91,
        '端州': 95,
        '鼎湖': 80,
        '封开': 83,
        '高要': 80,
        '云城': 75,
        '怀集': 55,
        '广宁': 50,
        '其他': 30
    },
    
    # 高亮配置
    '高亮配置': {
        '软件工程': '9C0006',  # 红色
        '计算机类': '9C5700'   # 淡黄棕色
    },
    
    # 评分样式配置
    '评分样式': {
        '优秀': {
            '分数线': 88,
            '背景色': 'C6EFCE',
            '字体色': '006100'
        },
        '良好': {
            '分数线': 80,
            '背景色': 'BCE094',
            '字体色': '006100'
        },
        '一般': {
            '分数线': 0,  # 低于80分
            '背景色': 'FFC7CE',
            '字体色': '9C0006'
        }
    },
    
    # 应届生样式配置
    '应届生样式': {
        '否': {
            '背景色': 'FFC7CE',
            '字体色': '9C0006'
        },
        '2025届高校毕业生': {
            '背景色': 'C6EFCE',
            '字体色': '006100'
        },
        '应届毕业生': {
            '背景色': 'BCE094',
            '字体色': '006100'
        }
    }
}

def load_data(filepath):
    """读取并合并所有sheet的数据"""
    excel_file = pd.ExcelFile(filepath)
    all_dfs = []
    
    # 遍历所有sheet并读取数据
    for sheet_name in excel_file.sheet_names:
        df = excel_file.parse(sheet_name, skiprows=1)  
        all_dfs.append(df)
    
    # 合并所有数据
    return pd.concat(all_dfs, ignore_index=True)

def calculate_score(row, config):
    """岗位评分算法"""
    score = 0
    
    # 适配度评分
    for key, value in config['专业适配度'].items():
        if key in row['本科专业\n名称及代码']:
            score += config['优先级权重']['适配度'] * value
            break
    
    # 竞争度评估（城市梯度）
    if row['考区'] in config['城市竞争度']:
        score += config['优先级权重']['城市竞争度'] * config['城市竞争度'][row['考区']]
    
    # 是否应届生
    if row['是否限应届毕业生报考'] in config['应届生状态']:
        score += config['优先级权重']['应届生'] * config['应届生状态'][row['是否限应届毕业生报考']]
    
    # 学历专业权重
    if row['学历'] in config['学历评分']:
        score += config['优先级权重']['学历'] * config['学历评分'][row['学历']]
 
    # 发展度评估（岗位层级）
    if '乡镇' in row['职位简介']:
        score += config['优先级权重']['发展度'] * config['岗位层级评分']['乡镇']
    else:
        for key in ['局', '政府', '院', '所']:
            if key in row['招考单位']:
                score += config['优先级权重']['发展度'] * config['岗位层级评分'][key]
                break
        else:
            score += config['优先级权重']['发展度'] * config['岗位层级评分']['其他']
    
    # 就职地
    for location, value in config['就职地区评分'].items():
        if location != '其他' and location in row['招考单位']:
            score += config['优先级权重']['就职地'] * value
            break
    else:
        score += config['优先级权重']['就职地'] * config['就职地区评分']['其他']
    
    return round(score, 1)

def highlight_keyword_in_cell(cell, keyword, font_color="FF0000"):
    """修改单元格内指定关键词的颜色"""
    original_text = str(cell.value)  # 确保转换为字符串
    target_index = original_text.find(keyword)
    
    if target_index == -1:
        return  # 关键词不存在时跳过
    
    # 分割文本
    before = original_text[:target_index]
    keyword_part = original_text[target_index:target_index+len(keyword)]
    after = original_text[target_index+len(keyword):]

    # 构建富文本块
    blocks = []
    if before:
        blocks.append(TextBlock(InlineFont(), before))
    blocks.append(TextBlock(InlineFont(color=font_color), keyword_part))
    if after:
        blocks.append(TextBlock(InlineFont(), after))

    # 重新赋值给单元格
    cell.value = CellRichText(*blocks)

def generate_report(df, config, output_path='岗位筛选报告.xlsx'):

    print(f"生成报告路径: {output_path}")  # 添加日志

    """生成带格式的Excel报告"""
    writer = pd.ExcelWriter(output_path, engine='openpyxl')
    df.to_excel(writer, index=False)
    
    wb = writer.book
    ws = wb.active
    
    # 评分样式设置
    score_styles = {}
    for level, style in config['评分样式'].items():
        score_styles[level] = {
            'fill': PatternFill(start_color=style['背景色'], end_color=style['背景色'], fill_type='solid'),
            'font': Font(color=style['字体色'])
        }
    
    # 应届生样式设置
    yj_styles = {}
    for status, style in config['应届生样式'].items():
        yj_styles[status] = {
            'fill': PatternFill(start_color=style['背景色'], end_color=style['背景色'], fill_type='solid'),
            'font': Font(color=style['字体色'])
        }
    
    # 获取列的索引
    score_column_index = df.columns.get_loc('岗位评分') + 1
    yingjie_index = df.columns.get_loc('是否限应届毕业生报考') + 1
    major_index = df.columns.get_loc('本科专业\n名称及代码') + 1
    
    for row in ws.iter_rows(min_row=2, max_col=len(df.columns), max_row=len(df) + 1):
        # 获取单元格
        score_cell = row[score_column_index - 1] 
        score_value = score_cell.value
        
        yj_cell = row[yingjie_index - 1]
        yj_value = yj_cell.value
        
        mj_cell = row[major_index - 1]
        
        # 高亮专业关键词
        for keyword, color in config['高亮配置'].items():
            highlight_keyword_in_cell(mj_cell, keyword, color)
        
        # 应用应届生单元格样式
        if yj_value in yj_styles:
            yj_cell.fill = yj_styles[yj_value]['fill']
            yj_cell.font = yj_styles[yj_value]['font']
        
        # 应用评分单元格样式
        if score_value is not None:
            if score_value >= config['评分样式']['优秀']['分数线']:
                score_cell.fill = score_styles['优秀']['fill']
                score_cell.font = score_styles['优秀']['font']
            elif score_value >= config['评分样式']['良好']['分数线']:
                score_cell.fill = score_styles['良好']['fill']
                score_cell.font = score_styles['良好']['font']
            else:
                score_cell.fill = score_styles['一般']['fill']
                score_cell.font = score_styles['一般']['font']
    
    writer.close()
    print(f"报告文件大小: {os.path.getsize(output_path)} bytes")  # 添加日志
    return output_path

def process_data(file_path, user_config=None):
    """处理数据的主函数"""
    # 合并默认配置和用户配置
    config = DEFAULT_CONFIG.copy()
    print("User Config:")
    print(user_config)
    if user_config:
        # 深度更新配置
        for key, value in user_config.items():
            if isinstance(value, dict) and key in config and isinstance(config[key], dict):
                config[key].update(value)
            else:
                config[key] = value
    
    # 读取数据
    raw_df = load_data(file_path)
    
    # 根据用户配置筛选数据
    filtered_df = raw_df[
        (raw_df['学历'].isin(config['学历要求'])) &
        (raw_df['考区'].isin(config['意向城市'])) &
        (raw_df['本科专业\n名称及代码'].str.contains('|'.join(config['专业关键词']), na=False)) &
        (~raw_df['其他要求'].fillna('').str.contains('|'.join(config['排除证书'])))
    ].copy()
    
    # 计算评分
    filtered_df['岗位评分'] = filtered_df.apply(lambda row: calculate_score(row, config), axis=1)
    
    # 排序结果
    result_df = filtered_df.sort_values(by='岗位评分', ascending=False)
    
    return result_df, config

@app.route('/api/process', methods=['POST'])
def api_process():
    # 检查是否上传了文件
    if 'file' not in request.files:
        return jsonify({"error": "没有上传文件"}), 400
    
    file = request.files['file']
    if file.filename == '':
        return jsonify({"error": "未选择文件"}), 400
    
    # 检查配置参数
    user_config = None
    if 'config' in request.form:
        try:
            user_config = json.loads(request.form['config'])
            print("Received user config:")
            print(user_config)
        except json.JSONDecodeError:
            return jsonify({"error": "配置格式错误"}), 400
    
    # 保存上传的文件
    temp_dir = tempfile.mkdtemp()
    input_path = os.path.join(temp_dir, file.filename)
    file.save(input_path)
    
    try:
        # 处理数据
        result_df, config = process_data(input_path, user_config)
        
        # 生成报告
        output_path = os.path.join(temp_dir, '岗位筛选报告.xlsx')
        report_path = generate_report(result_df, config, output_path)
        
        # 返回生成的报告

        # 检查文件是否存在
        if not os.path.exists(report_path):
            raise FileNotFoundError("生成的报告文件不存在")

        return send_file(
            report_path,
            as_attachment=True,
            download_name='岗位筛选报告.xlsx',
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'  # 关键修复
        )
    
    except Exception as e:
        return jsonify({"error": str(e)}), 500
    finally:
        # 清理临时文件
        import shutil
        shutil.rmtree(temp_dir, ignore_errors=True)

@app.route('/api/config/default', methods=['GET'])
def get_default_config():
    return jsonify(DEFAULT_CONFIG)

if __name__ == '__main__':
    app.run(debug=True, host='0.0.0.0', port=5000)